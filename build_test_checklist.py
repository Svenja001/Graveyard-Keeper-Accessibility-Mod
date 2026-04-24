"""Build an Excel test checklist for the unreleased changelog entries across all GYK mods."""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT_PATH = r"F:\OneDrive\Development\Game-Mods\Graveyard Keeper\reports\GYK_Test_Checklist.xlsx"

# Each tuple: (Mod, Version, Area, Test Step, Expected Result)
CHECKLIST = [
    # AddStraightToTable 2.4.11
    ("Add Straight To Table", "2.4.11", "Update Notifier", "Launch game with a newer version of this mod live on Nexus; reach main menu.", "Main-menu update notice lists Add Straight To Table with its newer version; clicking opens its Nexus page."),
    ("Add Straight To Table", "2.4.11", "Config UI", "Toggle the new 'Check for Updates' setting off, restart, reach main menu.", "Update notice is suppressed for this mod."),

    # AlchemyResearchRedux 0.1.5
    ("Alchemy Research Redux", "0.1.5", "Update Notifier", "Reach main menu with this mod installed and a newer version live on Nexus.", "Mod appears in the main-menu update notice and its entry opens the Nexus page when clicked."),
    ("Alchemy Research Redux", "0.1.5", "Config UI", "Open BepInEx Configuration Manager with a pre-existing config file from the previous version.", "Section headers now use `── Name ──` style and previously set values are preserved unchanged."),

    # AppleTreesEnhanced 2.7.13
    ("Apple Trees Enhanced", "2.7.13", "Localisation", "Set game language to Simplified Chinese; load save.", "All mod-provided text appears translated (no raw English fallback)."),
    ("Apple Trees Enhanced", "2.7.13", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Entry for this mod shown in the update notice; click opens Nexus page."),
    ("Apple Trees Enhanced", "2.7.13", "Config UI", "Open Configuration Manager; check every option description.", "Sections use `── Name ──` style; every setting has a readable description; existing values preserved."),

    # AutoLootHeavies 3.5.5
    ("Auto-Loot Heavies", "3.5.5", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated (no English fallback)."),
    ("Auto-Loot Heavies", "3.5.5", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice; clicking opens its Nexus page."),
    ("Auto-Loot Heavies", "3.5.5", "Config UI", "Open Configuration Manager with an existing config.", "Sections use `── Name ──` style, every option has a description, existing values preserved."),

    # BeamMeUpGerry 3.1.5
    ("Beam Me Up Gerry", "3.1.5", "Bug Fix — Area Gating", "Start a new-ish save, teleport to a zone (e.g. quarry) you haven't walked into on foot yet, then try to build an area-gated structure like the quarry mining hut.", "Area is unlocked post-teleport; gated buildings are placeable immediately without first walking in."),
    ("Beam Me Up Gerry", "3.1.5", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated (no English fallback)."),
    ("Beam Me Up Gerry", "3.1.5", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in the update notice; click opens Nexus."),
    ("Beam Me Up Gerry", "3.1.5", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style, sub-options indented under parent toggle, every entry has a description, existing values preserved."),

    # BringOutYerDead 0.2.6
    ("Bring Out Yer Dead", "0.2.6", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Bring Out Yer Dead", "0.2.6", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in the update notice; click opens Nexus."),
    ("Bring Out Yer Dead", "0.2.6", "Config UI", "Open Configuration Manager in Normal mode (not Advanced).", "Debug Logging option is visible without the Advanced filter."),
    ("Bring Out Yer Dead", "0.2.6", "Config UI", "Review section names and descriptions.", "Section names are clearer, descriptions are fuller."),
    ("Bring Out Yer Dead", "0.2.6", "Debug Logging", "Enable Debug Logging, play normally for a bit, review BepInEx log.", "Log output is less spammy than before and useful for bug reports."),

    # DecompDelight 0.1.7
    ("Decomp Delight", "0.1.7", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice; click opens Nexus page."),
    ("Decomp Delight", "0.1.7", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style, every colour entry has a description, existing values preserved."),

    # EconomyReloaded 1.3.10
    ("Economy Reloaded", "1.3.10", "Bug Fix — Vendor Tier", "Disable both inflation and deflation in settings. Load save, buy and sell items at a vendor until normally the level-up bar would advance.", "Vendor tier progress bar advances with every buy/sell even when prices are locked."),
    ("Economy Reloaded", "1.3.10", "Regression — Restock", "Play through a full in-game day with this mod enabled; observe shop restock at end of day.", "Shops restock at vanilla cadence — no aggressive over-restocking compared to base game."),
    ("Economy Reloaded", "1.3.10", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice; click opens Nexus page."),

    # Exhaustless 3.5.4
    ("Exhaustless", "3.5.4", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Exhaustless", "3.5.4", "Update Notifier", "Reach main menu with a newer version live on Nexus; also disable new 'Check for Updates' setting.", "Mod entry shows in notice when enabled; hidden when disabled."),

    # FasterCraftReloaded 1.4.11
    ("Faster Craft Reloaded", "1.4.11", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Faster Craft Reloaded", "1.4.11", "Bug Fix — Repair Crafts", "Set a high Craft Speed Multiplier (e.g. x10). Start a repair craft on a broken soul machine or damaged building.", "Repair completes normally without exhausting the player's entire energy bar per frame; formerly impossible repairs are now possible."),
    ("Faster Craft Reloaded", "1.4.11", "Config UI", "Open Configuration Manager.", "Section headings clearer, every option has a description. Speed multipliers indented under their parent toggle."),
    ("Faster Craft Reloaded", "1.4.11", "Debug Logging", "Enable Debug Logging, run multiple crafts; review log.", "One log line per craft session (not per frame). One-time in-game dialog appeared on load reminding Debug Logging is on, translated to current language."),
    ("Faster Craft Reloaded", "1.4.11", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears; click opens Nexus."),
    ("Faster Craft Reloaded", "1.4.11", "Config Descriptions", "Read the Craft Speed Multiplier and Speed Up Zombie Mines descriptions.", "Multiplier description clarifies gardens/composting/zombie stations have their own toggles. Zombie Mines description references stone yard, marble deposit, iron mine (not zombie mill)."),

    # FogBeGone 3.4.12
    ("Fog Be Gone", "3.4.12", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice; click opens Nexus."),

    # GerrysJunkTrunk 1.9.7
    ("Gerry's Junk Trunk", "1.9.7", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Gerry's Junk Trunk", "1.9.7", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Gerry's Junk Trunk", "1.9.7", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style; every option has a description; existing values preserved."),

    # GetOuttaMaWay 0.1.5
    ("Get Outta Ma Way", "0.1.5", "Feature — Drop Placement", "Chop a tree, break a rock, mine a block with this mod enabled (toggle ON).", "Logs/stones/blocks drop next to the tree/rock instead of flying to the player's feet."),
    ("Get Outta Ma Way", "0.1.5", "Feature — Drop Placement toggle off", "Disable the new toggle and repeat the chop/break/mine.", "Vanilla behaviour restored — items fly to feet again."),
    ("Get Outta Ma Way", "0.1.5", "Feature — Grace Window", "Chop a tree while standing on a work spot; observe the dropped log's physics for ~1.5s.", "Fresh drop cannot push the player off the spot within the grace window; pushing works again afterwards."),
    ("Get Outta Ma Way", "0.1.5", "Feature — Grace Window slider", "Set the grace-window slider to 0.25s, 1.5s, and 5s; test each.", "Grace duration respects the slider value at all three extremes."),
    ("Get Outta Ma Way", "0.1.5", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Get Outta Ma Way", "0.1.5", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style; existing values preserved."),

    # GiveMeMoar 1.2.14
    ("Give Me Moar", "1.2.14", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Give Me Moar", "1.2.14", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Give Me Moar", "1.2.14", "Config UI — Migration", "Launch with a pre-existing config file using the old '01.' / '02.' section names.", "Sections now use '── N. Name ──' style; all multiplier values carry over; nothing reset."),
    ("Give Me Moar", "1.2.14", "Bug Fix — Multiply Sticks", "Enable Multiply Sticks; chop trees and observe stick drops. Disable it; chop again.", "Sticks multiplied when ON; vanilla stick counts when OFF. Default is ON on fresh config."),
    ("Give Me Moar", "1.2.14", "Feature — Resource Multiplier coverage", "Set Resource Multiplier to e.g. x3; farm wheat/cabbage/carrot/beet/onion/lentils/pumpkin/hops/hemp/grapes, mine marble/diamonds, chop logs (billets/planks/beams/flitches), collect peat/salt/water/alcohol/eggs/milk/metal scrap; kill an enemy; harvest common body parts (blood/flesh/fat/skin/bone/skull).", "All listed items multiplied. Organs and specialised body parts remain at vanilla drops."),
    ("Give Me Moar", "1.2.14", "Feature — Categories", "In the new Categories section, toggle Seeds ON but Crops OFF; harvest.", "Only seeds get the multiplier; crops stay vanilla. Repeat logic for each of the nine category toggles to sanity-check."),
    ("Give Me Moar", "1.2.14", "Feature — Category defaults", "On fresh install, inspect the Categories defaults.", "Crops/Bugs/Ores/Logs/Enemy Drops ON by default; Seeds/Misc/Body Parts OFF."),
    ("Give Me Moar", "1.2.14", "Feature — Craft Output Multiplier", "Set Craft Output Multiplier to x5. Craft billets from logs at the sawhorse and planks at the carpenter's bench.", "Craft yields are multiplied (e.g. 1 log → 5 billets)."),
    ("Give Me Moar", "1.2.14", "Feature — Per-station overrides", "Set Per-station overrides to `mf_sawhorse_1=10;mf_anvil_2=3`. Craft at those stations and at another station.", "Overridden stations use per-station values; other stations fall back to the global multiplier."),
    ("Give Me Moar", "1.2.14", "Feature — Progression exclusions", "Attempt station upgrade, tier upgrade, repair craft, object placement, refugee build-desk craft, a tool craft, a weapon craft, an armour craft, a sermon scroll.", "All of these produce vanilla quantities (not multiplied) by default."),
    ("Give Me Moar", "1.2.14", "Feature — Exclusion toggles", "Turn off the progression exclusion toggle; retry a station upgrade.", "Progression craft is now multiplied. Turn back on — reverts."),
    ("Give Me Moar", "1.2.14", "Feature — Excluded Craft IDs", "Add a specific craft ID to the 'Excluded Craft IDs' field; craft it.", "That craft is excluded from the multiplier."),
    ("Give Me Moar", "1.2.14", "Feature — Multiplier range", "Set a multiplier below 1 (e.g. 0.5) and run a harvest/craft.", "Outputs are reduced below vanilla. Values at 0.1 and 50 both accepted."),
    ("Give Me Moar", "1.2.14", "Debug Logging Reminder", "Enable Debug Logging, restart game, load save.", "One-time in-game reminder pops up in the current language."),
    ("Give Me Moar", "1.2.14", "Debug Logging Tags", "Play normally with Debug Logging on; review BepInEx log.", "Log lines tagged with [Drop], [Faith], [TechPoints], [CraftApply] contexts."),

    # GraveChangesRedux 0.1.8
    ("Grave Changes Redux", "0.1.8", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Grave Changes Redux", "0.1.8", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style; every option has a clearer description; existing values preserved."),

    # IBuildWhereIWant 1.7.11
    ("I Build Where I Want", "1.7.11", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("I Build Where I Want", "1.7.11", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("I Build Where I Want", "1.7.11", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style; every option has a description; existing values preserved."),

    # INeedSticks 1.6.10
    ("I Need Sticks", "1.6.10", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("I Need Sticks", "1.6.10", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # KeepersCandles 0.1.8
    ("Keepers Candles", "0.1.8", "Feature — Incense", "Place an incense burner in the world.", "It auto-lights on placement and burns forever (no consumption)."),
    ("Keepers Candles", "0.1.8", "Feature — Incense Keybind", "Set a keybind/controller button for 'Extinguish Incense'; stand near a lit incense and press it.", "Nearest lit incense is extinguished and the unused incense item is returned."),
    ("Keepers Candles", "0.1.8", "Feature — Keybind Separation", "Bind incense key + keep candle key bound; have both lit near the player; press candle key only.", "Only candles are extinguished — incense untouched. Vice versa for incense key."),
    ("Keepers Candles", "0.1.8", "Config UI — Migration", "Load with an existing config that had a custom candle keybind.", "Candle keybind appears under the renamed 'Extinguish Candle Keybind' entry; user's custom value preserved."),
    ("Keepers Candles", "0.1.8", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Keepers Candles", "0.1.8", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Keepers Candles", "0.1.8", "Config UI — Advanced", "Open Configuration Manager in Normal mode.", "Debug Logging is visible without toggling Advanced."),
    ("Keepers Candles", "0.1.8", "Config UI — Sections", "Open Configuration Manager.", "New sections present: Advanced / Candles & Incenses / Church / Controls; player-facing descriptions on each option."),
    ("Keepers Candles", "0.1.8", "Config Migration", "Load with a config from the previous version.", "Keybinds, distance, column toggle all migrated without loss."),
    ("Keepers Candles", "0.1.8", "Localisation Files", "Check mod folder `lang/` directory exists and contains JSON translation files.", "Translations are loaded from `lang/*.json`, editable externally."),
    ("Keepers Candles", "0.1.8", "Debug Logging Reminder", "Enable Debug Logging, restart, load save.", "One-time in-game reminder appears."),

    # LongerDays 1.6.10
    ("Longer Days", "1.6.10", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # MaxButtonsRedux 1.3.12
    ("Max Buttons Redux", "1.3.12", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Max Buttons Redux", "1.3.12", "Dependency Check", "Uninstall Rest In Patches and try to launch with this mod only.", "Mod refuses to load / reports the missing dependency cleanly."),
    ("Max Buttons Redux", "1.3.12", "Bug Fix — Max with Star-Quality Ingredients", "Open a craft that needs star-quality ingredients (e.g. armor, cooked meal) in COLLAPSED mode; press Max button.", "Max button sets the correct craft amount (not 1)."),

    # MiscBitsAndBobs 2.3.5
    ("Misc Bits And Bobs", "2.3.5", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Misc Bits And Bobs", "2.3.5", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Misc Bits And Bobs", "2.3.5", "Config UI", "Open Configuration Manager.", "Every toggle has a plain-English description; Player/Porter speed sliders indent under their parent toggle."),
    ("Misc Bits And Bobs", "2.3.5", "Config Migration", "Load with an existing config.", "Sections re-organised as Advanced / Audio / UI / Gameplay / Movement / Church / Misc; existing values migrated."),
    ("Misc Bits And Bobs", "2.3.5", "Debug Logging", "Turn on Debug Logging; trigger letterbox suppression, intro skip, prayer removal, tavern oven fuel change, zombie pyre input, Halloween event, sermon line substitution, Sprint Reloaded detection, and church visitor eviction.", "Each action produces its own labeled log line. Disabling logging yields no overhead."),

    # NewGameAtBottom 2.2.9
    ("New Game At Bottom", "2.2.9", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # NoIntros 2.2.9
    ("No Intros", "2.2.9", "Update Notifier", "Reach main menu with a newer version live on Nexus; also test 'Check for Updates' setting toggle.", "Mod entry appears when enabled, hidden when disabled."),

    # NoTimeForFishing 3.2.9
    ("No Time For Fishing", "3.2.9", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # PrayTheDayAway 0.3.6
    ("Pray The Day Away", "0.3.6", "Localisation — Chinese", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Pray The Day Away", "0.3.6", "Localisation — Fallback Fix", "Set game language to any non-English (e.g. French); load save.", "Mod text translated to selected language — not English fallback."),
    ("Pray The Day Away", "0.3.6", "Localisation — Live Language Switch", "Change in-game language mid-session.", "Mod text updates immediately without restart."),
    ("Pray The Day Away", "0.3.6", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Pray The Day Away", "0.3.6", "Config UI", "Open Configuration Manager.", "Sections use `── Name ──` style; sub-options indented; every entry has a description; existing values preserved."),
    ("Pray The Day Away", "0.3.6", "Config UI — Advanced Order", "Open Configuration Manager in Normal mode.", "Advanced section is at the TOP of the list; Debug Logging always visible."),
    ("Pray The Day Away", "0.3.6", "Debug Logging Reminder", "Enable Debug Logging, restart, load save.", "One-time in-game dialog warns you Debug is on."),

    # QueueEverything 2.1.13
    ("Queue Everything", "2.1.13", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Queue Everything", "2.1.13", "Bug Fix — Build Menu", "Enable Auto Max Normal Crafts and Auto Max Multi-Quality Crafts. Try to place graves, garden beds, and other build-menu items.", "Build-menu items remain placeable (single-unit behaviour) — they do not become stuck at an unplaceable quantity."),
    ("Queue Everything", "2.1.13", "Bug Fix — Controller Focus", "With a controller connected, set 'Auto Select Craft Button With Controller' = ON. Open a multi-quality craft (e.g. carrot soup, any i/ii/iii recipe) and expand it.", "Focus lands on the big 'Craft Now' button (not the first ingredient). Pressing A/Cross crafts immediately."),
    ("Queue Everything", "2.1.13", "Bug Fix — Controller Focus OFF", "Same setup but toggle = OFF; re-expand a multi-quality craft.", "Focus lands on the first ingredient (vanilla behaviour)."),
    ("Queue Everything", "2.1.13", "Bug Fix — Controller Focus Non-Multi-Quality", "With controller + toggle ON, expand a non-multi-quality craft.", "No errors; no misplaced focus; toggle has no visible effect (expected — button does not exist for these)."),
    ("Queue Everything", "2.1.13", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Queue Everything", "2.1.13", "Dependency Check", "Uninstall Rest In Patches; launch with this mod only.", "Mod refuses to load and reports missing dependency."),

    # RegenerationReloaded 1.1.11
    ("Regeneration Reloaded", "1.1.11", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # RestInPatches 0.1.2
    ("Rest In Patches", "0.1.2", "Performance — Drops", "Load a zone with lots of ground drops; walk past and around them.", "Noticeably smoother frame time vs. previous build; no visible highlight flickering."),
    ("Rest In Patches", "0.1.2", "Performance — HUD", "Play for a day cycle and observe the HUD (day number, RGB weather values, hint label).", "HUD values update only when they change — watch for no visible per-frame refresh artifacts."),
    ("Rest In Patches", "0.1.2", "Bug Check — Player Tint", "Trigger player tint changes (buffs, weather, day/night).", "Tint overlay updates correctly when the colour actually changes; no stuck tints."),
    ("Rest In Patches", "0.1.2", "Performance — World Objects", "Roam a zone with many idle objects; cross zone transitions; enter build mode.", "No regression in frame pacing; build-mode entry/exit feels responsive; pathfinding still works."),
    ("Rest In Patches", "0.1.2", "Config UI", "Open Configuration Manager.", "Section headers now use `── Name ──` style; existing values preserved."),

    # SaveNow 2.5.12
    ("Save Now", "2.5.12", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Save Now", "2.5.12", "Bug Fix — Teleport Stone", "Enter a dungeon, use a teleport stone to leave.", "Saving is enabled immediately after teleporting out (not stuck blocked)."),
    ("Save Now", "2.5.12", "Bug Fix — Controller Save Button", "With controller, open tech tree, trade screen, and quantity-transfer screen. Press LT (the manual-save button).", "Manual save does NOT trigger while any menu is open — LT continues to navigate within the menu."),
    ("Save Now", "2.5.12", "Bug Check — Controller Save Normal", "Close all menus; press the manual-save controller button.", "Save fires as expected when no menu is open."),
    ("Save Now", "2.5.12", "Update Notifier", "Reach main menu with a newer version live on Nexus; also test 'Check for Updates' setting toggle.", "Mod entry appears when enabled, hidden when disabled."),

    # ShowMeMoar 0.1.11
    ("Show Me Moar", "0.1.11", "High-DPI Prompt", "On a Windows box with display scaling > 100%, first-launch with this mod installed.", "Main-menu prompt offers to apply the high-DPI fix automatically."),
    ("Show Me Moar", "0.1.11", "High-DPI Apply", "Click Yes on the prompt, restart the game.", "Main menu renders sharply at native resolution (not blurry). If still blurry, open Settings → Display and change resolution once — the README dialog should say so."),
    ("Show Me Moar", "0.1.11", "High-DPI Undo", "Open Configuration Manager → '── 9. High-DPI Fix ──' section; undo the fix.", "Fix is reverted; registry override removed."),
    ("Show Me Moar", "0.1.11", "High-DPI Silence", "In the same section, silence the prompt.", "Prompt no longer appears on next launch."),
    ("Show Me Moar", "0.1.11", "High-DPI Translations", "Repeat the prompt flow in each of DE, EN, ES, FR, IT, JA, KO, PL, PT-BR, RU, ZH-CN.", "Prompt, status, and result dialogs all appear in the chosen language."),
    ("Show Me Moar", "0.1.11", "Docs Check", "Re-read the mod-page manual workaround steps and the in-game success dialog.", "Mod page says 'High DPI scaling override' = 'Application' (not 'System'). Success dialog mentions the Settings → Display resolution trick. No mention of the Zoom slider workaround."),
    ("Show Me Moar", "0.1.11", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # TheSeedEqualizer 1.3.11
    ("The Seed Equalizer", "1.3.11", "Localisation", "Set game language to Simplified Chinese; load save.", "Mod text is translated."),
    ("The Seed Equalizer", "1.3.11", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("The Seed Equalizer", "1.3.11", "Config UI — Advanced", "Open Configuration Manager in Normal mode.", "Debug Logging visible without the Advanced filter."),
    ("The Seed Equalizer", "1.3.11", "Config UI — Sections", "Open Configuration Manager.", "Sections: Advanced / Player Gardens / Zombie Gardens / Refugee Gardens / Waste / All Gardens. Every option has a player-facing description."),
    ("The Seed Equalizer", "1.3.11", "Config Migration", "Load with an existing config from the previous version.", "Old '01. / 02. / 03.' sections are rewritten; values carry across."),
    ("The Seed Equalizer", "1.3.11", "Debug Logging", "Turn on Debug Logging; play through a garden cycle.", "Log records: which gardens modified, seeds added, crafts that got crop waste, and reasons any craft was skipped."),
    ("The Seed Equalizer", "1.3.11", "Debug Logging Reminder", "With Debug on, load a save and wait until morning.", "Start-of-day on-screen reminder appears in current language (test a few from the 11 supported)."),

    # ThoughtfulReminders 2.2.11
    ("Thoughtful Reminders", "2.2.11", "Feature — Wake-Up Delay", "Go to sleep. Set Wake-Up Delay to default (a couple of seconds).", "Reminder message appears AFTER the sleep screen fades — not hidden behind it."),
    ("Thoughtful Reminders", "2.2.11", "Feature — Wake-Up Delay 0", "Set Wake-Up Delay to 0; sleep.", "Reminder appears instantly on wake-up."),
    ("Thoughtful Reminders", "2.2.11", "Localisation — Chinese", "Set language to Simplified Chinese; load save.", "Reminder text translated."),
    ("Thoughtful Reminders", "2.2.11", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Thoughtful Reminders", "2.2.11", "Localisation — Fallback Fix", "Set a non-English language; load save.", "Reminder translated to that language (no English fallback)."),
    ("Thoughtful Reminders", "2.2.11", "Localisation — Live Switch", "Change language mid-session.", "Reminder picks up the new language without restart."),
    ("Thoughtful Reminders", "2.2.11", "Config UI", "Open Configuration Manager.", "Sections: Advanced / Reminders. Clearer descriptions on each option."),
    ("Thoughtful Reminders", "2.2.11", "Debug Logging Reminder", "Enable Debug Logging, restart, load save.", "One-time in-game reminder appears."),

    # TreesNoMore 2.5.12
    ("Trees No More", "2.5.12", "Localisation", "Set language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Trees No More", "2.5.12", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Trees No More", "2.5.12", "Config UI", "Open Configuration Manager.", "Sections reorganised with full player-facing descriptions; existing values migrated."),
    ("Trees No More", "2.5.12", "Debug Logging", "Turn on Debug Logging; chop a tree; review log.", "Every step of the chop-handling flow is traceable in the log."),

    # WheresMaPoints 0.3.4
    ("Wheres Ma Points", "0.3.4", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),

    # WheresMaStorage 2.1.14
    ("Wheres Ma Storage", "2.1.14", "Localisation", "Set language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Wheres Ma Storage", "2.1.14", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Collect Drops On Load (Pockets mode)", "Set 'Collect Drops On Game Load' to pockets mode; reload save with loose loot on the ground.", "Loose loot pulled straight into inventory on load."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Collect Drops On Load (Pile mode)", "Switch to pile-by-house mode; reload save with loose loot scattered around (including large items).", "Loot piled next to house instead of pocketed; large items are included in the sweep."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Near-House Dump Zone Radius", "Set slider to 8 tiles (default). Drop items near house, save and reload.", "Items already within 8 tiles of house stay put — not re-shuffled. Adjust slider to 3 and 16 to confirm radius respected."),
    ("Wheres Ma Storage", "2.1.14", "Performance — Vendor Windows", "Enable 'Show Only Personal Inventory'; open a vendor.", "Vendor window opens noticeably faster than prior version."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Vendor Widget Hides", "Toggle Hide Stockpile/Tavern/Soul/Warehouse Shop Widgets; open a vendor.", "Vendor window respects the toggles (hidden widgets don't appear)."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Inventory Dimming inverted", "Toggle 'Inventory Dimming' ON.", "Greyed-out style actually applies (previously inverted)."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Vendor Dimming on full list", "At a vendor with the full shared list showing, inspect non-personal items.", "Non-personal items are NOT dimmed."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Loot Magnet Range", "Set Player Loot Magnet Range slider to 1.8, 10, 20 tiles; drop items at various distances.", "Drops within the set radius are sucked in; beyond are not."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Bag Widget Columns", "Open a bag widget.", "Layout is 5-column (matching other widgets), not the old 3-column."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Quest Items Not Swept", "Load a save with a quest item dropped on the ground.", "Load-time sweep leaves the quest item untouched (not pocketed or piled)."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Split Inventory Sliders", "Check Configuration Manager for the two separate sliders: player inventory and container inventory.", "Both sliders present, independent. Old single-slider value auto-migrated to both on first launch."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Slider Minimum 0", "Set either slider to 0.", "Accepted; no extra slots granted (vanilla behaviour)."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Slider Resizes Existing Containers", "Build a chest with slider at +20; lower slider to +5.", "Existing chest now reports +5 capacity (not stuck at +20)."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Shrink Prompt Yes", "Fill player inventory; lower slider below current item count. Click Yes on the prompt.", "Excess items drop at player's feet; container shrinks accordingly. For chest test: excess items drop next to the chest."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Shrink Prompt No", "Repeat previous test but click No.", "Slider rolls back; nothing is lost."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Hidden-Item Recovery", "Load a save that was affected by the silent-shrink bug in prior versions.", "Previously hidden items re-appear next time the affected container is opened."),
    ("Wheres Ma Storage", "2.1.14", "Feature — Config Manager Auto-close on Shrink", "Trigger the shrink prompt from within the Configuration Manager window.", "Config Manager closes automatically so the Yes/No prompt is unobstructed."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Modify Inventory Size OFF", "Turn off Modify Inventory Size; play and build/destroy chests, pick up drops on load.", "Player inventory remains at 20 slots throughout; does not snap back to bigger size."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Modify Inventory Size ON", "Turn it back on; play for a while.", "Player inventory stays at the slider's bigger size; extra slots never silently disappear."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Live Apply", "While in-game, change slider or toggle master on/off.", "Change applies immediately (no need to interact with anything first)."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Quarry/Mill Containers", "At the quarry (or zombie mill), craft from elsewhere first with Exclude options on, then return to the quarry/mill to craft.", "Quarry/zombie mill containers are still visible (previously vanished until build/destroy)."),
    ("Wheres Ma Storage", "2.1.14", "Bug Fix — Empty Rows Toggle", "Turn OFF 'Don't Show Empty Rows In Inventory' while in personal inventory.", "Empty rows now appear in personal inventory (not just chests)."),
    ("Wheres Ma Storage", "2.1.14", "Performance — Slider Drag", "Drag any slider in Configuration Manager quickly.", "No in-game hitching during drag; change is applied at most once per frame."),
    ("Wheres Ma Storage", "2.1.14", "Config UI", "Open Configuration Manager.", "Parent toggles followed by `└` indented dependents; `── Name ──` section style; existing values preserved across the rename."),
    ("Wheres Ma Storage", "2.1.14", "Diagnostics Log", "Without enabling Debug Logging, play for a while and inspect BepInEx log.", "Inventory check timings always logged."),

    # WheresMaVeggies 0.1.9
    ("Wheres Ma Veggies", "0.1.9", "Localisation", "Set language to Simplified Chinese; load save.", "Mod text is translated."),
    ("Wheres Ma Veggies", "0.1.9", "Update Notifier", "Reach main menu with a newer version live on Nexus.", "Mod entry appears in update notice."),
    ("Wheres Ma Veggies", "0.1.9", "Config UI — Migration", "Load with an existing config using old numbered section names.", "Section names no longer have leading numbers; existing values preserved."),
]


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Checklist"

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    mod_fill = PatternFill("solid", fgColor="D9E1F2")
    mod_font = Font(bold=True, size=11)

    headers = ["Mod", "Version", "Area", "Test Step", "Expected Result", "Pass", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    row = 2
    last_mod = None
    for mod, version, area, step, expected in CHECKLIST:
        if mod != last_mod:
            cell = ws.cell(row=row, column=1, value=f"{mod} — v{version}")
            cell.fill = mod_fill
            cell.font = mod_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = border
            row += 1
            last_mod = mod

        values = [mod, version, area, step, expected, "", ""]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.alignment = Alignment(
                horizontal="center" if col == 6 else "left",
                vertical="top",
                wrap_text=True,
            )
            cell.border = border
        row += 1

    widths = {1: 24, 2: 10, 3: 22, 4: 56, 5: 56, 6: 8, 7: 28}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 110

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Total test rows: {sum(1 for _ in CHECKLIST)}")


if __name__ == "__main__":
    build_workbook()
